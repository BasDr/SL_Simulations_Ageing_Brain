import numpy as np
import matplotlib.pyplot as plt
import pandas as pd
import os
from StuartLandauModel import simulate_SL
from numba import njit
import concurrent.futures
import scipy.signal as signal
import scipy.stats as stat
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import matplotlib.colors as mcolors
from mpl_toolkits.mplot3d import Axes3D

os.chdir('/vol/specs07/BasDrost/SL_simulation/')

# Define the grid points
gc = np.logspace(np.log10(1), np.log10(70), 40)  # Logarithmic Global Coupling
md = np.linspace(0, 20, 21)  # Mean Delay
a_range = [-50, -20, -10, -5, -1, 0, 1, 5, 10, 20, 50]   # Alpha Peak Frequency

# Reduce the number of y-axis ticks for better visibility
num_ticks = 8
tick_positions = np.linspace(0, len(gc) - 1, num_ticks).astype(int)
tick_labels = [f'{gc[i]:.2f}' for i in tick_positions]

#figure size
old_vs_young_x = 15
old_vs_young_y = len(a_range)*5+len(a_range)*2
simulation_only_x = len(a_range)*5+len(a_range)*2
simulation_only_y = 7

amount_of_a = len(a_range)

all_max_corr_Y = []
all_max_corr_E = []
all_mean_corr_Y = []
all_mean_corr_E = []
all_variance_corr_Y = []
all_variance_corr_E = []

all_psd_peak = []
all_variance_psd_peak = []

all_mean_psd_power = []
all_var_psd_power = []

all_mean_meta = []
all_var_meta = []

all_mean_sync = []
all_var_sync = []

all_mean_SE = []
all_var_SE = []

for a_val in a_range:
    Max_corr_Y_list = np.zeros((len(gc), len(md)))
    Max_corr_E_list = np.zeros((len(gc), len(md)))
    Mean_corr_Y_list = np.zeros((len(gc), len(md)))
    Mean_corr_E_list = np.zeros((len(gc), len(md)))
    variance_mean_corr_Y_list = np.zeros((len(gc), len(md)))
    variance_mean_corr_E_list = np.zeros((len(gc), len(md)))

    psd_peak_list= np.zeros((len(gc), len(md)))
    variance_psd_peak_list = np.zeros((len(gc), len(md)))    

    mean_psd_power = np.zeros((len(gc), len(md)))
    var_psd_power = np.zeros((len(gc), len(md)))

    mean_meta = np.zeros((len(gc), len(md)))
    var_meta = np.zeros((len(gc), len(md)))

    mean_sync = np.zeros((len(gc), len(md)))
    var_sync = np.zeros((len(gc), len(md)))

    mean_SE = np.zeros((len(gc), len(md)))
    var_SE = np.zeros((len(gc), len(md)))

    for i,gc_val in enumerate(gc):
        for j,md_val in enumerate(md):

            try:
                data = np.load('Mean_simulation_results/SL_simulation_results_C{:.2f}_md{:.1f}_a{:.2f}.npz'.format(gc_val, md_val, a_val))
                result_Y = data['corr_info_Y']
                result_E = data['corr_info_E']
                mean_psd_peak_list = data['psd_peak']
                batch_psd_power = data['psd_power']
                batch_meta = data['meta_info']
                batch_sync = data['sync_info']
                batch_SE = data['SE_info']

            except FileNotFoundError:
                result_Y = [np.nan]
                result_E = [np.nan]
                mean_psd_peak_list = [np.nan]
                batch_psd_power = [np.nan]
                batch_meta = [np.nan]
                batch_sync = [np.nan]
                batch_SE = [np.nan]
                continue
            

            psd_peak_list[i][j] = np.mean(mean_psd_peak_list)
            variance_psd_peak_list[i][j] = np.var(mean_psd_peak_list)

            mean_psd_power[i][j] = np.mean(batch_psd_power)
            mean_meta[i][j] = np.mean(batch_meta)
            mean_sync[i][j] = np.mean(batch_sync)
            mean_SE[i][j] = np.mean(batch_SE)

            var_psd_power[i][j] = np.var(batch_psd_power)
            var_meta[i][j] = np.var(batch_meta)
            var_sync[i][j] = np.var(batch_sync)
            var_SE[i][j] = np.var(batch_SE)


            Max_corr_Y_list[i][j] = np.max(result_Y)
            Max_corr_E_list[i][j] = np.max(result_E)

            Mean_corr_Y_list[i][j] = np.mean(result_Y)
            Mean_corr_E_list[i][j] = np.mean(result_E)

            variance_mean_corr_Y_list[i][j] = np.var(result_Y)
            variance_mean_corr_E_list[i][j] = np.var(result_E)
            
            
    
    all_max_corr_Y.append(Max_corr_Y_list)
    all_max_corr_E.append(Max_corr_E_list)
    all_mean_corr_Y.append(Mean_corr_Y_list)
    all_mean_corr_E.append(Mean_corr_E_list)
    all_psd_peak.append(psd_peak_list)
    all_variance_psd_peak.append(variance_psd_peak_list)
    all_variance_corr_Y.append(variance_mean_corr_Y_list)
    all_variance_corr_E.append(variance_mean_corr_E_list)

    all_mean_psd_power.append(mean_psd_power)
    all_var_psd_power.append(var_psd_power)
    all_mean_meta.append(mean_meta)
    all_var_meta.append(var_meta)
    all_mean_sync.append(mean_sync)
    all_var_sync.append(var_sync)
    all_mean_SE.append(mean_SE)
    all_var_SE.append(var_SE)

def plot_mean_correlation(filter=bool, contour_lines=bool):
    # Plot Mean Correlation
    a = 0
    plt.figure(figsize=(old_vs_young_x, old_vs_young_y))
    for a_val in range(len(a_range)):
        plot_list = np.nan * np.zeros((len(gc), len(md)))
        for i in range(len(gc)):
            for j in range(len(md)):
                if filter:
                    if 8 < all_psd_peak[a_val][i][j] < 13:
                        plot_list[i][j] = all_mean_corr_Y[a_val][i][j]
                else:
                    if all_mean_corr_Y[a_val][i][j] > 0.36:
                        print(f"Plot - a: {a_range[a_val]}, gc: {gc[i]}, md: {md[j]}, alpha: {all_psd_peak[a_val][i][j]}")
                        print(i)
                        print(j)
                    plot_list[i][j] = all_mean_corr_Y[a_val][i][j]
        a += 1

        # Plot Mean Correlation for Youth
        plt.subplot(amount_of_a, 2, a)
        plt.imshow(plot_list, cmap='viridis', aspect='auto', origin='lower', 
                   #vmin=0.1, vmax=0.35
                   )
        if contour_lines:
            step_x_low, step_y_low, step_x_high, step_y_high = return_contour(a_val)
            plt.step(step_x_low, step_y_low, where='pre', color='red', linewidth=3, label='Step function')
            plt.step(step_x_high, step_y_high, where='pre', color='red', linewidth=3, label='Step function')
        plt.colorbar()
        plt.title(f'Mean Correlation Youth for a={a_range[a_val]}')
        plt.xlabel('Mean Delay')
        plt.ylabel('Global Coupling')
        plt.xticks(ticks=np.linspace(0, len(md) - 1, len(md)), labels=[f'{md[i]:.0f}' for i in range(len(md))])
        plt.yticks(ticks=tick_positions, labels=tick_labels)

        plot_list = np.nan * np.zeros((len(gc), len(md)))
        for i in range(len(gc)):
            for j in range(len(md)):
                if filter:
                    if 8 < all_psd_peak[a_val][i][j] < 13:
                        plot_list[i][j] = all_mean_corr_E[a_val][i][j]
                else:
                    plot_list[i][j] = all_mean_corr_E[a_val][i][j]
        a += 1

        # Plot Mean Correlation for Elder
        plt.subplot(amount_of_a, 2, a)
        plt.imshow(plot_list, cmap='viridis', aspect='auto', origin='lower', 
                   #vmin=0.1, vmax=0.5
                   )
        if contour_lines:
            step_x_low, step_y_low, step_x_high, step_y_high = return_contour(a_val)
            plt.step(step_x_low, step_y_low, where='pre', color='red', linewidth=3, label='Step function')
            plt.step(step_x_high, step_y_high, where='pre', color='red', linewidth=3, label='Step function')
        plt.colorbar()
        plt.title(f'Mean Correlation Elder for a={a_range[a_val]}')
        plt.xlabel('Mean Delay')
        plt.ylabel('Global Coupling')
        plt.xticks(ticks=np.linspace(0, len(md) - 1, len(md)), labels=[f'{md[i]:.0f}' for i in range(len(md))])
        plt.yticks(ticks=tick_positions, labels=tick_labels)

    plt.tight_layout()
    if filter:
        plt.savefig('graphs_mean/1_to_70/Mean_Correlation_new_filtered.png')
    elif contour_lines:
        plt.savefig('graphs_mean/1_to_70/Mean_Correlation_new_contour.png')
    else:
        plt.savefig('graphs_mean/1_to_70/Mean_Correlation_new_whole.png')

def plot_max_correlation(filter=bool, contour_lines=bool):

    a = 0
    plt.figure(figsize=(old_vs_young_x, old_vs_young_y))
    for a_val in range(len(a_range)):
        plot_list = np.nan * np.zeros((len(gc), len(md)))
        for i in range(len(gc)):
            for j in range(len(md)):
                if filter:
                    if 8 < all_psd_peak[a_val][i][j] < 13:
                        plot_list[i][j] = all_max_corr_Y[a_val][i][j]
                else:
                    plot_list[i][j] = all_max_corr_Y[a_val][i][j]


        a += 1

        # Plot Max Correlation for Youth
        plt.subplot(amount_of_a, 2, a)
        plt.imshow(plot_list, cmap='viridis', aspect='auto', origin='lower', 
                   #vmin=0, vmax=0.5
                   )  # Adjusted origin
        if contour_lines:
            step_x_low, step_y_low, step_x_high, step_y_high = return_contour(a_val)
            plt.step(step_x_low, step_y_low, where='pre', color='red', linewidth=3, label='Step function')
            plt.step(step_x_high, step_y_high, where='pre', color='red', linewidth=3, label='Step function')
        plt.colorbar()
        plt.title(f'Max Correlation Youth for a={a_range[a_val]}')
        plt.xlabel('Mean Delay')
        plt.ylabel('Global Coupling')
        plt.xticks(ticks=np.linspace(0, len(md) - 1, len(md)), labels=[f'{md[i]:.0f}' for i in range(len(md))])
        plt.yticks(ticks=tick_positions, labels=tick_labels)

        plot_list = np.nan * np.zeros((len(gc), len(md)))
        for i in range(len(gc)):
            for j in range(len(md)):
                if filter:
                    if 8 < all_psd_peak[a_val][i][j] < 13:
                        plot_list[i][j] = all_max_corr_E[a_val][i][j]
                else:
                    plot_list[i][j] = all_max_corr_E[a_val][i][j]
        a += 1

        # Plot Max Correlation for Elder
        plt.subplot(amount_of_a, 2, a)
        plt.imshow(plot_list, cmap='viridis', aspect='auto', origin='lower', 
                   #vmin=0, vmax=0.5
                   )  # Adjusted origin
        if contour_lines:
            step_x_low, step_y_low, step_x_high, step_y_high = return_contour(a_val)
            plt.step(step_x_low, step_y_low, where='pre', color='red', linewidth=3, label='Step function')
            plt.step(step_x_high, step_y_high, where='pre', color='red', linewidth=3, label='Step function')
        plt.colorbar()
        plt.title(f'Max Correlation Elder for a={a_range[a_val]}')
        plt.xlabel('Mean Delay')
        plt.ylabel('Global Coupling')
        plt.xticks(ticks=np.linspace(0, len(md) - 1, len(md)), labels=[f'{md[i]:.0f}' for i in range(len(md))])
        plt.yticks(ticks=tick_positions, labels=tick_labels)

    plt.tight_layout()
    if filter:
        plt.savefig('graphs_mean/1_to_70/Max_Correlation_new_filtered.png')
    elif contour_lines:
        plt.savefig('graphs_mean/1_to_70/Max_Correlation_new_contour.png')
    else:
        plt.savefig('graphs_mean/1_to_70/Max_Correlation_new_whole.png')


def plot_psd_peak(filter=bool):
    a = 0
    plt.figure(figsize=(simulation_only_x, simulation_only_y))
    for a_val in range(len(a_range)):
        plot_list = np.nan * np.zeros((len(gc), len(md)))
        for i in range(len(gc)):
            for j in range(len(md)):
                if filter:
                    if 8 < all_psd_peak[a_val][i][j] < 13:
                        plot_list[i][j] = all_psd_peak[a_val][i][j]
                else:
                    plot_list[i][j] = all_psd_peak[a_val][i][j]
        a += 1

        # Plot PSD Peak
        plt.subplot(1, amount_of_a, a)
        plt.imshow(plot_list, cmap='viridis', aspect='auto', origin='lower')  # Adjusted origin
        plt.colorbar()
        plt.title(f'PSD Peak for a={a_range[a_val]}')
        plt.xlabel('Mean Delay')
        plt.ylabel('Global Coupling')
        plt.xticks(ticks=np.linspace(0, len(md) - 1, len(md)), labels=[f'{md[i]:.0f}' for i in range(len(md))])
        plt.yticks(ticks=tick_positions, labels=tick_labels)

    plt.tight_layout()
    if filter:
        plt.savefig('graphs_mean/1_to_70/PSD_Peak_filtered.png')
    else:
        plt.savefig('graphs_mean/1_to_70/PSD_Peak.png')

def plot_SEM_peak(peak_filter=bool, contour_lines=bool):

    a=0
    plt.figure(figsize=(old_vs_young_x, old_vs_young_y/2+10))
    for a_val in range(len(a_range)):
        
        if peak_filter:
            plot_list = np.nan * np.zeros((len(gc), len(md)))
            for i in range(len(gc)):
                for j in range(len(md)):
                    if 8 < all_psd_peak[a_val][i][j] < 13:
                        #plot SEM/mean (for signifcant value should be <0.05)
                        plot_list[i][j] = np.sqrt(all_variance_psd_peak[a_val][i][j]/len(mean_psd_peak_list))/all_psd_peak[a_val][i][j]
        else:
            plot_list = np.sqrt(all_variance_psd_peak[a_val]/len(mean_psd_peak_list))/all_psd_peak[a_val]
        a += 1

        # Plot Mean Correlation for Youth
        plt.subplot(amount_of_a//2+1, 2, a)
        plt.imshow(plot_list, cmap='viridis', aspect='auto', origin='lower', 
                vmin=0, vmax=0.05
                )
        if contour_lines:
            step_x_low, step_y_low, step_x_high, step_y_high = return_contour(a_val)
            plt.step(step_x_low, step_y_low, where='pre', color='red', linewidth=3, label='Step function')
            plt.step(step_x_high, step_y_high, where='pre', color='red', linewidth=3, label='Step function')
        plt.colorbar()
        plt.title(f'SEM/mean Youth  for a={a_range[a_val]}')
        plt.xlabel('Mean Delay')
        plt.ylabel('Global Coupling')
        plt.xticks(ticks=np.linspace(0, len(md) - 1, len(md)), labels=[f'{md[i]:.0f}' for i in range(len(md))])
        plt.yticks(ticks=tick_positions, labels=tick_labels)

    plt.tight_layout()
    if peak_filter:
        plt.savefig('graphs_mean/1_to_70/SEM_on_psd_peak_filtered.png')
    elif contour_lines:
        plt.savefig('graphs_mean/1_to_70/SEM_on_psd_peak_contour.png')
    else:
        plt.savefig('graphs_mean/1_to_70/SEM_on_psd_peak_whole.png')
    

def plot_SEM_corr(peak_filter=bool, contour_lines=bool):
    
    a=0
    plt.figure(figsize=(old_vs_young_x, old_vs_young_y))
    for a_val in range(len(a_range)):
        
        if peak_filter:
            plot_list = np.nan * np.zeros((len(gc), len(md)))
            for i in range(len(gc)):
                for j in range(len(md)):
                    if 8 < all_psd_peak[a_val][i][j] < 13:
                        #plot SEM/mean (for signifcant value should be <0.05)
                        plot_list[i][j] = np.sqrt(all_variance_corr_Y[a_val][i][j]/len(mean_psd_peak_list))/all_mean_corr_Y[a_val][i][j]
        else:
            plot_list = np.sqrt(all_variance_corr_Y[a_val]/len(mean_psd_peak_list))/all_mean_corr_Y[a_val]
        a += 1

        # Plot Mean Correlation for Youth
        plt.subplot(amount_of_a, 2, a)
        plt.imshow(plot_list, cmap='viridis', aspect='auto', origin='lower', 
                vmin=0, vmax=0.05
                )
        if contour_lines:
            step_x_low, step_y_low, step_x_high, step_y_high = return_contour(a_val)
            plt.step(step_x_low, step_y_low, where='pre', color='red', linewidth=3, label='Step function')
            plt.step(step_x_high, step_y_high, where='pre', color='red', linewidth=3, label='Step function')
        plt.colorbar()
        plt.title(f'SEM/mean Youth  for a={a_range[a_val]}')
        plt.xlabel('Mean Delay')
        plt.ylabel('Global Coupling')
        plt.xticks(ticks=np.linspace(0, len(md) - 1, len(md)), labels=[f'{md[i]:.0f}' for i in range(len(md))])
        plt.yticks(ticks=tick_positions, labels=tick_labels)
        

        if peak_filter:
            plot_list = np.nan * np.zeros((len(gc), len(md)))
            for i in range(len(gc)):
                for j in range(len(md)):
                    if 8 < all_psd_peak[a_val][i][j] < 13:
                        #plot SEM/mean (for signifcant value should be <0.05)
                        plot_list[i][j] = np.sqrt(all_variance_corr_E[a_val][i][j]/len(mean_psd_peak_list))/all_mean_corr_E[a_val][i][j]
        else:
            plot_list = np.sqrt(all_variance_corr_E[a_val]/len(mean_psd_peak_list))/all_mean_corr_E[a_val]
    
        a += 1

        # Plot Mean Correlation for Elder
        plt.subplot(amount_of_a, 2, a)
        plt.imshow(plot_list, cmap='viridis', aspect='auto', origin='lower', 
                vmin=0, vmax=0.05
                )
        if contour_lines:
            step_x_low, step_y_low, step_x_high, step_y_high = return_contour(a_val)
            plt.step(step_x_low, step_y_low, where='pre', color='red', linewidth=3, label='Step function')
            plt.step(step_x_high, step_y_high, where='pre', color='red', linewidth=3, label='Step function')
        plt.colorbar()
        plt.title(f'SEM/mean Elder  for a={a_range[a_val]}')
        plt.xlabel('Mean Delay')
        plt.ylabel('Global Coupling')
        plt.xticks(ticks=np.linspace(0, len(md) - 1, len(md)), labels=[f'{md[i]:.0f}' for i in range(len(md))])
        plt.yticks(ticks=tick_positions, labels=tick_labels)

    plt.tight_layout()
    if peak_filter:
        plt.savefig('graphs_mean/1_to_70/SEM_on_mean_corr_filtered.png')
    elif contour_lines:
        plt.savefig('graphs_mean/1_to_70/SEM_on_mean_corr_contour.png')
    else:
        plt.savefig('graphs_mean/1_to_70/SEM_on_mean_corr_whole.png')

def top_10_mean_corr():
    # Create a workbook and worksheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Top 10 Mean Correlations"
    
    # Define header row
    header = ["Type", "Score", "a", "gc", "md", "Alpha Peak Freq", "Alpha power", "Meta", "Sync", "SE"]
    fill_color = PatternFill(start_color="F7923F", end_color="F7923F", fill_type="solid")
    
    # Write data
    row_number = 1
    coll_ofsett= 1
    sheet.cell(row=row_number, column=coll_ofsett, value=header[0])
    sheet.cell(row=row_number, column=coll_ofsett + 1, value=header[1])
    sheet.cell(row=row_number, column=coll_ofsett + 2, value=header[2])
    sheet.cell(row=row_number, column=coll_ofsett + 3, value=header[3])
    sheet.cell(row=row_number, column=coll_ofsett + 4, value=header[4])
    sheet.cell(row=row_number, column=coll_ofsett + 5, value=header[5])
    sheet.cell(row=row_number, column=coll_ofsett + 6, value=header[6])
    sheet.cell(row=row_number, column=coll_ofsett + 7, value=header[7])
    sheet.cell(row=row_number, column=coll_ofsett + 8, value=header[8])
    sheet.cell(row=row_number, column=coll_ofsett + 9, value=header[9])

    
    for col in range(0, len(header)):
        cell = sheet.cell(row=row_number, column=coll_ofsett+col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    row_number += 1

    for a_val in range(len(a_range)):
        #if a_val%2==0 and a_val!=0:
            #coll_ofsett+=len(header)+1
            #row_number=1
        
        

        # Process for Y values
        list = np.zeros((len(gc), len(md)))
        for i in range(len(gc)):
            for j in range(len(md)):
                if True: #7 < all_psd_peak[a_val][i][j] < 13:
                    list[i][j] = all_mean_corr_Y[a_val][i][j]
        
        # Get top 10 scores and their indices
        flat_indices = np.argsort(list, axis=None)[-30:]  # Indices of the top 10 scores
        top_scores = list.flatten()[flat_indices]
        top_scores= top_scores[::-1]         # Corresponding scores
        flat_indices= flat_indices[::-1]   ### INDICES MUST MATCH SCORES
        top_indices = np.unravel_index(flat_indices, list.shape)  # Convert to 2D indices
        
        top10=0
        for score, i, j in zip(top_scores, top_indices[0], top_indices[1]):
            if 7 < all_psd_peak[a_val][i][j] < 13:
                if top10==10:
                    #sheet.append([""])
                    #row_number += 1
                    break
                else:
                    sheet.cell(row=row_number, column=coll_ofsett, value="Y")
                    sheet.cell(row=row_number, column=coll_ofsett + 1, value=f"{round(score, 3)}")# ± {round(np.sqrt(all_variance_corr_Y[a_val][i][j] / len(mean_psd_peak_list)), 3)}")
                    sheet.cell(row=row_number, column=coll_ofsett + 2, value=a_range[a_val])
                    sheet.cell(row=row_number, column=coll_ofsett + 3, value=round(gc[i], 2))
                    sheet.cell(row=row_number, column=coll_ofsett + 4, value=int(md[j]))
                    sheet.cell(row=row_number, column=coll_ofsett + 5, value=f"{round(all_psd_peak[a_val][i][j], 1)}")# ± {round(np.sqrt(all_variance_psd_peak[a_val][i][j] / len(mean_psd_peak_list)), 3)}")
                    sheet.cell(row=row_number, column=coll_ofsett + 6, value=f"{round(all_mean_psd_power[a_val][i][j], 1)}")# ± {round(np.sqrt(all_var_psd_power[a_val][i][j] / len(mean_psd_peak_list)), 3)}")
                    sheet.cell(row=row_number, column=coll_ofsett + 7, value=f"{round(all_mean_meta[a_val][i][j], 2)}")# ± {round(np.sqrt(all_var_meta[a_val][i][j] / len(mean_psd_peak_list)), 3)}")
                    sheet.cell(row=row_number, column=coll_ofsett + 8, value=f"{round(all_mean_sync[a_val][i][j], 2)}")# ± {round(np.sqrt(all_var_sync[a_val][i][j] / len(mean_psd_peak_list)), 3)}")
                    sheet.cell(row=row_number, column=coll_ofsett + 9, value=f"{round(all_mean_SE[a_val][i][j], 2)}")# ± {round(np.sqrt(all_var_SE[a_val][i][j] / len(mean_psd_peak_list)), 3)}")
                    if 9.6<= round(all_psd_peak[a_val][i][j], 1) <=12:
                        for col in range(0, len(header)):
                            sheet.cell(row=row_number, column=coll_ofsett + col).fill = fill_color
                    row_number += 1
                    top10+=1
                
        
        # Process for E values
        top_list = np.zeros((len(gc), len(md)))
        for i in range(len(gc)):
            for j in range(len(md)):
                if True: #7 < all_psd_peak[a_val][i][j] < 13:
                    top_list[i][j] = all_mean_corr_E[a_val][i][j]
        
        # Get top 10 scores and their indices
        flat_indices = np.argsort(top_list, axis=None)[-30:]
        top_scores = top_list.flatten()[flat_indices]
        top_scores = top_scores[::-1]
        flat_indices= flat_indices[::-1]   ### INDICES MUST MATCH SCORES
        top_indices = np.unravel_index(flat_indices, top_list.shape)

        # Write each top score with its corresponding data
        top10=0
        for score, i, j in zip(top_scores, top_indices[0], top_indices[1]):
            if 7 < all_psd_peak[a_val][i][j] < 13:
                if top10==10:
                    #sheet.append([""])
                    #row_number += 1
                    break
                else:
                    sheet.cell(row=row_number, column=coll_ofsett, value="E")
                    sheet.cell(row=row_number, column=coll_ofsett + 1, value=f"{round(score, 3)}")# ± {round(np.sqrt(all_variance_corr_E[a_val][i][j] / len(mean_psd_peak_list)), 3)}")
                    sheet.cell(row=row_number, column=coll_ofsett + 2, value=a_range[a_val])
                    sheet.cell(row=row_number, column=coll_ofsett + 3, value=round(gc[i], 2))
                    sheet.cell(row=row_number, column=coll_ofsett + 4, value=int(md[j]))
                    sheet.cell(row=row_number, column=coll_ofsett + 5, value=f"{round(all_psd_peak[a_val][i][j], 1)}")# ± {round(np.sqrt(all_variance_psd_peak[a_val][i][j] / len(mean_psd_peak_list)), 3)}")
                    sheet.cell(row=row_number, column=coll_ofsett + 6, value=f"{round(all_mean_psd_power[a_val][i][j], 1)}")# ± {round(np.sqrt(all_var_psd_power[a_val][i][j] / len(mean_psd_peak_list)), 3)}")
                    sheet.cell(row=row_number, column=coll_ofsett + 7, value=f"{round(all_mean_meta[a_val][i][j], 2)}")# ± {round(np.sqrt(all_var_meta[a_val][i][j] / len(mean_psd_peak_list)), 3)}")
                    sheet.cell(row=row_number, column=coll_ofsett + 8, value=f"{round(all_mean_sync[a_val][i][j], 2)}")# ± {round(np.sqrt(all_var_sync[a_val][i][j] / len(mean_psd_peak_list)), 3)}")
                    sheet.cell(row=row_number, column=coll_ofsett + 9, value=f"{round(all_mean_SE[a_val][i][j], 2)}")# ± {round(np.sqrt(all_var_SE[a_val][i][j] / len(mean_psd_peak_list)), 3)}")
                    if 7.8 <= round(all_psd_peak[a_val][i][j], 1) <= 8.5:
                        for col in range(0, len(header)):
                            sheet.cell(row=row_number, column=coll_ofsett + col).fill = fill_color
                    row_number += 1
                    top10+=1
                    
    # Auto-adjust column widths
    for col in sheet.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        sheet.column_dimensions[col_letter].width = max_length + 2

    # Save the workbook
    workbook.save("/vol/specs07/BasDrost/SL_simulation/graphs_mean/1_to_70/Mean_analyses_formatted_error.xlsx")

def top_10_mean_corr_all():
    # Create a workbook and worksheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Top 10 Mean Correlations"
    
    # Define header row
    header = ["Type", "Score", "a", "gc", "md", "Alpha Peak Freq", "Alpha power", "Meta", "Sync", "SE"]
    fill_color = PatternFill(start_color="F7923F", end_color="F7923F", fill_type="solid")
    
    # Write data
    row_number = 1
    coll_ofsett= 1
    sheet.cell(row=row_number, column=coll_ofsett, value=header[0])
    sheet.cell(row=row_number, column=coll_ofsett + 1, value=header[1])
    sheet.cell(row=row_number, column=coll_ofsett + 2, value=header[2])
    sheet.cell(row=row_number, column=coll_ofsett + 3, value=header[3])
    sheet.cell(row=row_number, column=coll_ofsett + 4, value=header[4])
    sheet.cell(row=row_number, column=coll_ofsett + 5, value=header[5])
    sheet.cell(row=row_number, column=coll_ofsett + 6, value=header[6])
    sheet.cell(row=row_number, column=coll_ofsett + 7, value=header[7])
    sheet.cell(row=row_number, column=coll_ofsett + 8, value=header[8])
    sheet.cell(row=row_number, column=coll_ofsett + 9, value=header[9])

    
    for col in range(0, len(header)):
        cell = sheet.cell(row=row_number, column=coll_ofsett+col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    row_number += 1

    for a_val in range(len(a_range)):
        #if a_val%2==0 and a_val!=0:
            #coll_ofsett+=len(header)+1
            #row_number=1
        
        

        # Process for Y values
        list = np.zeros((len(gc), len(md)))
        for i in range(len(gc)):
            for j in range(len(md)):
                if True: #7 < all_psd_peak[a_val][i][j] < 13:
                    list[i][j] = all_mean_corr_Y[a_val][i][j]
        
        # Get top 10 scores and their indices
        flat_indices = np.argsort(all_mean_corr_Y[a_val], axis=None)[-30:]  # Indices of the top 10 scores
        top_scores = all_mean_corr_Y[a_val].flatten()[flat_indices]
        top_scores= top_scores[::-1]         # Corresponding scores
        flat_indices= flat_indices[::-1]   ### INDICES MUST MATCH SCORES
        top_indices = np.unravel_index(flat_indices, all_mean_corr_Y[a_val].shape)  # Convert to 2D indices
        
        top10=0
        for score, i, j in zip(top_scores, top_indices[0], top_indices[1]):
            if 7 < all_psd_peak[a_val][i][j] < 13:
                if top10==10:
                    #sheet.append([""])
                    #row_number += 1
                    break
                else:
                    try:
                        data = np.load('Mean_simulation_results/SL_simulation_results_C{:.2f}_md{:.1f}_a{:.2f}.npz'.format(gc[i], md[j], a_range[a_val]))
                        result_Y = data['corr_info_Y']
                        mean_psd_peak_list = data['psd_peak']
                        batch_psd_power = data['psd_power']
                        batch_meta = data['meta_info']
                        batch_sync = data['sync_info']
                        batch_SE = data['SE_info']
                    except FileNotFoundError:
                        result_Y = [np.nan]
                        mean_psd_peak_list = [np.nan]
                        batch_psd_power = [np.nan]
                        batch_meta = [np.nan]
                        batch_sync = [np.nan]
                        batch_SE = [np.nan]
                        continue
                    for z in range(len(result_Y)):
                        sheet.cell(row=row_number, column=coll_ofsett, value="Y")
                        sheet.cell(row=row_number, column=coll_ofsett + 1, value=f"{round(result_Y[z], 3)}")
                        sheet.cell(row=row_number, column=coll_ofsett + 2, value=a_range[a_val])
                        sheet.cell(row=row_number, column=coll_ofsett + 3, value=round(gc[i], 2))
                        sheet.cell(row=row_number, column=coll_ofsett + 4, value=int(md[j]))
                        sheet.cell(row=row_number, column=coll_ofsett + 5, value=f"{round(mean_psd_peak_list[z], 1)}")
                        sheet.cell(row=row_number, column=coll_ofsett + 6, value=f"{round(batch_psd_power[z], 1)}")
                        sheet.cell(row=row_number, column=coll_ofsett + 7, value=f"{round(batch_meta[z], 2)}")
                        sheet.cell(row=row_number, column=coll_ofsett + 8, value=f"{round(batch_sync[z], 2)}")
                        sheet.cell(row=row_number, column=coll_ofsett + 9, value=f"{round(batch_SE[z], 2)}")
                        if 9.6<= round(mean_psd_peak_list[z], 1) <=12:
                            for col in range(0, len(header)):
                                sheet.cell(row=row_number, column=coll_ofsett + col).fill = fill_color
                        row_number += 1
                    top10+=1
                
        
        # Process for E values
        top_list = np.zeros((len(gc), len(md)))
        for i in range(len(gc)):
            for j in range(len(md)):
                if True: #7 < all_psd_peak[a_val][i][j] < 13:
                    top_list[i][j] = all_mean_corr_E[a_val][i][j]
        
        # Get top 10 scores and their indices
        flat_indices = np.argsort(top_list, axis=None)[-30:]
        top_scores = top_list.flatten()[flat_indices]
        top_scores = top_scores[::-1]
        flat_indices= flat_indices[::-1]   ### INDICES MUST MATCH SCORES
        top_indices = np.unravel_index(flat_indices, top_list.shape)

        # Write each top score with its corresponding data
        top10=0
        for score, i, j in zip(top_scores, top_indices[0], top_indices[1]):
            if 7 < all_psd_peak[a_val][i][j] < 13:
                if top10==10:
                    #sheet.append([""])
                    #row_number += 1
                    break
                else:
                    try:
                        data = np.load('Mean_simulation_results/SL_simulation_results_C{:.2f}_md{:.1f}_a{:.2f}.npz'.format(gc[i], md[j], a_range[a_val]))
                        result_E = data['corr_info_E']
                        mean_psd_peak_list = data['psd_peak']
                        batch_psd_power = data['psd_power']
                        batch_meta = data['meta_info']
                        batch_sync = data['sync_info']
                        batch_SE = data['SE_info']
                    except FileNotFoundError:
                        result_E = [np.nan]
                        mean_psd_peak_list = [np.nan]
                        batch_psd_power = [np.nan]
                        batch_meta = [np.nan]
                        batch_sync = [np.nan]
                        batch_SE = [np.nan]
                        continue
                    for z in range(len(result_Y)):
                        sheet.cell(row=row_number, column=coll_ofsett, value="E")
                        sheet.cell(row=row_number, column=coll_ofsett + 1, value=f"{round(result_E[z], 3)}")# ± {round(np.sqrt(all_variance_corr_E[a_val][i][j] / len(mean_psd_peak_list)), 3)}")
                        sheet.cell(row=row_number, column=coll_ofsett + 2, value=a_range[a_val])
                        sheet.cell(row=row_number, column=coll_ofsett + 3, value=round(gc[i], 2))
                        sheet.cell(row=row_number, column=coll_ofsett + 4, value=int(md[j]))
                        sheet.cell(row=row_number, column=coll_ofsett + 5, value=f"{round(mean_psd_peak_list[z], 1)}")
                        sheet.cell(row=row_number, column=coll_ofsett + 6, value=f"{round(batch_psd_power[z], 1)}")
                        sheet.cell(row=row_number, column=coll_ofsett + 7, value=f"{round(batch_meta[z], 2)}")
                        sheet.cell(row=row_number, column=coll_ofsett + 8, value=f"{round(batch_sync[z], 2)}")
                        sheet.cell(row=row_number, column=coll_ofsett + 9, value=f"{round(batch_SE[z], 2)}")
                        if 7.8 <= round(mean_psd_peak_list[z], 1) <= 8.5:
                            for col in range(0, len(header)):
                                sheet.cell(row=row_number, column=coll_ofsett + col).fill = fill_color
                        row_number += 1
                    top10+=1
                    
    # Auto-adjust column widths
    for col in sheet.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        sheet.column_dimensions[col_letter].width = max_length + 2

    # Save the workbook
    workbook.save("/vol/specs07/BasDrost/SL_simulation/graphs_mean/1_to_70/Mean_analyses_formatted_XXL.xlsx")

def plot_psd_parameter_relation(contour_lines=bool):
    a = 0
    plt.figure(figsize=(simulation_only_x, simulation_only_y))
    
    for a_val in range(len(a_range)):
        plt.subplot(1, amount_of_a, a+1)
        unique_colors = plt.cm.get_cmap('gist_ncar', 40)  # Generates 40 unique colors
        #norm = mcolors.BoundaryNorm(boundaries=np.arange(41)-0.5, ncolors=40)
        # Plot Mean Correlation for Youth
        plot_list = np.nan * np.zeros((len(gc), len(md)))
        
        for r in range(43):
            for i in range(len(gc)):
                for j in range(len(md)):
                    if r-0.5 <= all_psd_peak[a_val][i][j] < r+0.5:
                        plot_list[i][j] = r
        
        
        plt.imshow(plot_list, aspect='auto', origin='lower', cmap=unique_colors, interpolation='none'
                )
        plt.colorbar()

        
        for i in range(plot_list.shape[0]):  # Iterate over rows
            #iterate over columns in reverse order to find the first step
            for j in range(plot_list.shape[1]-1, -1, -1):  # Iterate over columns in reverse order
                value = int(plot_list[i, j])  # Get the value at this position
                plt.text(j, i, f"{value}", color='black', ha='center', va='center', 
                        fontsize=8, fontweight='bold')
        # turn step_x and step_y around
        if contour_lines:
            step_x_low, step_y_low, step_x_high, step_y_high = return_contour(a_val)
            plt.step(step_x_low, step_y_low, where='pre', color='red', linewidth=3, label='Step function')
            plt.step(step_x_high, step_y_high, where='pre', color='red', linewidth=3, label='Step function')

        plt.title(f'PSD peaks for a={a_range[a_val]}')  # Example title
        plt.xlabel('Mean Delay')
        plt.ylabel('Global Coupling')
        plt.xticks(ticks=np.linspace(0, plot_list.shape[1] - 1, plot_list.shape[1]), 
                labels=[f'{i:.0f}' for i in range(plot_list.shape[1])])
        plt.yticks(ticks=tick_positions, labels=tick_labels)
        a += 1
    plt.tight_layout()
    if contour_lines:
        plt.savefig('graphs_mean/1_to_70/PSD_Peak_color_contour.png')
    else:
        plt.savefig('graphs_mean/1_to_70/PSD_Peak_color.png')

def plot_psd_parameter_relation_dif_color():
    a = 0
    plt.figure(figsize=(simulation_only_x, simulation_only_y))
    
    for a_val in range(len(a_range)):
        plt.subplot(1, amount_of_a, a+1)
        base_colors = plt.colormaps.get_cmap('tab20').colors  # Get 20 distinct colors
        color_list = list(base_colors) * 2  # Repeat them to get 40 colors
        np.random.seed(20)
        np.random.shuffle(color_list)
        # Step 2: Manually reassign colors to avoid direct neighbors sharing colors
        final_colors = [None] * 40
        final_colors[0] = color_list[0]

        for i in range(1, 40):
            # Ensure the color is not the same as the previous one
            possible_colors = [c for c in color_list if not np.array_equal(c, final_colors[i - 1])]
            final_colors[i] = possible_colors[0]  # Pick the first valid one
            for idx, c in enumerate(color_list):
                if np.array_equal(c, possible_colors[0]):
                    del color_list[idx]  # Remove first occurrence
                    break  # Stop after removing one

        # Step 3: Create a colormap from these 40 assigned colors
        cmap = mcolors.ListedColormap(final_colors)
        #norm = mcolors.BoundaryNorm(boundaries=np.arange(41)-0.5, ncolors=40)
        # Plot Mean Correlation for Youth
        plot_list = np.nan * np.zeros((len(gc), len(md)))
        
        for r in range(43):
            for i in range(len(gc)):
                for j in range(len(md)):
                    if r-0.5 <= all_psd_peak[a_val][i][j] < r+0.5:
                        plot_list[i][j] = r
        
        plt.imshow(plot_list, aspect='auto', origin='lower', cmap=cmap, interpolation='nearest'
                )
        plt.colorbar()

        # Add contour lines at each integer level to show clear separations
        #plt.contour(plot_list, levels=np.arange(0, 41, 1), colors='black', linewidths=0.5)

        for i in range(plot_list.shape[0]):  # Iterate over rows
            for j in range(plot_list.shape[1]):  # Iterate over columns
                value = int(plot_list[i, j])  # Get the value at this position
                #text_color = cmap(norm(value))  # Get corresponding color from colormap
                plt.text(j, i, f"{value}", color='black', ha='center', va='center', 
                        fontsize=8, fontweight='bold')

        plt.title(f'PSD peaks for a={a_range[a_val]}')  # Example title
        plt.xlabel('Mean Delay')
        plt.ylabel('Global Coupling')
        plt.xticks(ticks=np.linspace(0, plot_list.shape[1] - 1, plot_list.shape[1]), 
                labels=[f'{i:.0f}' for i in range(plot_list.shape[1])])
        plt.yticks(ticks=tick_positions, labels=tick_labels)
        a += 1
    plt.tight_layout()
    plt.savefig('graphs_mean/1_to_70/PSD_Peak_random_color.png')

def add_scatter_fit(ax, x, y, color, label):
    """Helper function to plot scatter points and fit line."""
    ax.scatter(x, y, color=color, alpha=0.5, label=label, edgecolors='k')  # Scatter points
    fit = np.poly1d(np.polyfit(x, y, deg=1))  # Quadratic fit
    ax.plot(x, fit(x), color=color, linestyle='--', linewidth=2)  # Fit line


def analyse_top_10():
    E_Top10_gc_all_a = []
    E_Top10_md_all_a = []
    E_Top10_peak_alpha_all_a = []
    E_score_all_a = []

    Y_Top10_gc_all_a = []
    Y_Top10_md_all_a = []
    Y_Top10_peak_alpha_all_a = []
    Y_score_all_a = []
    for a_val in range(len(a_range)):
        E_Top10_gc=[]
        E_Top10_md=[]
        E_Top10_peak_alpha=[]
        E_Top10_score=[]

        Y_Top10_gc=[]
        Y_Top10_md=[]
        Y_Top10_peak_alpha=[]
        Y_Top10_score=[]

        E_top_list = np.zeros((len(gc), len(md)))
        Y_top_list = np.zeros((len(gc), len(md)))
        for i in range(len(gc)):
            for j in range(len(md)):
                if True: #7 < all_psd_peak[a_val][i][j] < 13:
                    E_top_list[i][j] = all_mean_corr_E[a_val][i][j]
                    Y_top_list[i][j] = all_mean_corr_Y[a_val][i][j]
        
        # Get top 10 scores and their indices
        flat_indices = np.argsort(E_top_list, axis=None)[-30:]
        top_scores = E_top_list.flatten()[flat_indices]
        top_scores = top_scores[::-1]
        flat_indices= flat_indices[::-1]   ### INDICES MUST MATCH SCORES
        E_top_indices = np.unravel_index(flat_indices, E_top_list.shape)   

        top10=0
        for score, i, j in zip(top_scores, E_top_indices[0], E_top_indices[1]):
            if 7 < all_psd_peak[a_val][i][j] < 13:
                if top10==10:
                    break
                else:
                    E_Top10_gc.append(gc[i])
                    E_Top10_md.append(md[j])
                    E_Top10_peak_alpha.append(all_psd_peak[a_val][i][j])
                    E_Top10_score.append(score)
                    top10+=1
        E_Top10_gc_all_a.append(E_Top10_gc)
        E_Top10_md_all_a.append(E_Top10_md)
        E_Top10_peak_alpha_all_a.append(E_Top10_peak_alpha)
        E_score_all_a.append(E_Top10_score)

        
        flat_indices = np.argsort(Y_top_list, axis=None)[-30:]
        top_scores = Y_top_list.flatten()[flat_indices]
        top_scores = top_scores[::-1]
        flat_indices= flat_indices[::-1]   ### INDICES MUST MATCH SCORES
        Y_top_indices = np.unravel_index(flat_indices, Y_top_list.shape)

        top10=0
        for score, i, j in zip(top_scores, Y_top_indices[0], Y_top_indices[1]):
            if 7 < all_psd_peak[a_val][i][j] < 13:
                if top10==10:
                    break
                else:
                    Y_Top10_gc.append(gc[i])
                    Y_Top10_md.append(md[j])
                    Y_Top10_peak_alpha.append(all_psd_peak[a_val][i][j])
                    Y_Top10_score.append(score)
                    top10+=1
        Y_Top10_gc_all_a.append(Y_Top10_gc)
        Y_Top10_md_all_a.append(Y_Top10_md)
        Y_Top10_peak_alpha_all_a.append(Y_Top10_peak_alpha)
        Y_score_all_a.append(Y_Top10_score)
        
    E_Top10_gc_mean = np.mean(E_Top10_gc_all_a, axis=1)
    E_Top10_md_mean = np.mean(E_Top10_md_all_a, axis=1)
    E_Top10_peak_alpha_mean = np.mean(E_Top10_peak_alpha_all_a, axis=1)
    E_Top10_score_mean = np.mean(E_score_all_a, axis=1)

    E_Top10_gc_var = np.var(E_Top10_gc_all_a, axis=1)
    E_Top10_md_var = np.var(E_Top10_md_all_a, axis=1)
    E_Top10_peak_alpha_var = np.var(E_Top10_peak_alpha_all_a, axis=1)
    E_Top10_score_var = np.var(E_score_all_a, axis=1)

    Y_Top10_gc_mean = np.mean(Y_Top10_gc_all_a, axis=1)
    Y_Top10_md_mean = np.mean(Y_Top10_md_all_a, axis=1)
    Y_Top10_peak_alpha_mean = np.mean(Y_Top10_peak_alpha_all_a, axis=1)
    Y_Top10_score_mean = np.mean(Y_score_all_a, axis=1)

    Y_Top10_gc_var = np.var(Y_Top10_gc_all_a, axis=1)
    Y_Top10_md_var = np.var(Y_Top10_md_all_a, axis=1)
    Y_Top10_peak_alpha_var = np.var(Y_Top10_peak_alpha_all_a, axis=1)
    Y_Top10_score_var = np.var(Y_score_all_a, axis=1)

    #plot all 4 parameters in one graph    
    fig, (ax1, ax5) = plt.subplots(1, 2, figsize=(16, 6), sharex=True)

    # --- First Dataset (Elder) ---
    ax1.set_xlabel("a_range")
    ax1.set_ylabel("GC Mean", color='b')
    ax1.plot(a_range, E_Top10_gc_mean, color='b', label="GC Mean")
    ax1.fill_between(a_range, E_Top10_gc_mean - np.sqrt(E_Top10_gc_var), E_Top10_gc_mean + np.sqrt(E_Top10_gc_var), color='b', alpha=0.2)
    #add_scatter_fit(ax1, a_range, E_Top10_gc_mean, 'b', "GC Mean")
    ax1.tick_params(axis='y', labelcolor='b')

    ax2 = ax1.twinx()
    ax2.set_ylabel("MD Mean", color='g')
    ax2.plot(a_range, E_Top10_md_mean, color='g', label="MD Mean")
    ax2.fill_between(a_range, E_Top10_md_mean - np.sqrt(E_Top10_md_var), E_Top10_md_mean + np.sqrt(E_Top10_md_var), color='g', alpha=0.2)
    #add_scatter_fit(ax2, a_range, E_Top10_md_mean, 'g', "MD Mean")
    ax2.tick_params(axis='y', labelcolor='g')

    ax3 = ax1.twinx()
    ax3.spines["right"].set_position(("outward", 60))
    ax3.set_ylabel("Peak Alpha Mean", color='r')
    ax3.plot(a_range, E_Top10_peak_alpha_mean, color='r', label="Peak Alpha Mean")
    ax3.fill_between(a_range, E_Top10_peak_alpha_mean - np.sqrt(E_Top10_peak_alpha_var), E_Top10_peak_alpha_mean + np.sqrt(E_Top10_peak_alpha_var), color='r', alpha=0.2)
    #add_scatter_fit(ax3, a_range, E_Top10_peak_alpha_mean, 'r', "Peak Alpha Mean")
    ax3.tick_params(axis='y', labelcolor='r')

    ax4 = ax1.twinx()
    ax4.spines["right"].set_position(("outward", 120))
    ax4.set_ylabel("Score", color='c')
    ax4.plot(a_range, E_Top10_score_mean, color='c', label="Score")
    ax4.fill_between(a_range, E_Top10_score_mean - np.sqrt(E_Top10_score_var), E_Top10_score_mean + np.sqrt(E_Top10_score_var), color='c', alpha=0.2)
    #add_scatter_fit(ax4, a_range, E_Top10_score_mean, 'c', "Score")
    ax4.tick_params(axis='y', labelcolor='c')

    ax1.grid()
    ax1.set_title("Elder Group: Multi-Axis Analysis with Fits")

    # --- Second Dataset (Youth) ---
    ax5.set_xlabel("a_range")
    ax5.set_ylabel("GC Mean", color='b')
    ax5.plot(a_range, Y_Top10_gc_mean, color='b', label="GC Mean")
    ax5.fill_between(a_range, Y_Top10_gc_mean - np.sqrt(Y_Top10_gc_var), Y_Top10_gc_mean + np.sqrt(Y_Top10_gc_var), color='b', alpha=0.2)
    #add_scatter_fit(ax5, a_range, Y_Top10_gc_mean, 'b', "GC Mean")
    ax5.tick_params(axis='y', labelcolor='b')

    ax6 = ax5.twinx()
    ax6.set_ylabel("MD Mean", color='g')
    ax6.plot(a_range, Y_Top10_md_mean, color='g', label="MD Mean")
    ax6.fill_between(a_range, Y_Top10_md_mean - np.sqrt(Y_Top10_md_var), Y_Top10_md_mean + np.sqrt(Y_Top10_md_var), color='g', alpha=0.2)
    #add_scatter_fit(ax6, a_range, Y_Top10_md_mean, 'g', "MD Mean")
    ax6.tick_params(axis='y', labelcolor='g')

    ax7 = ax5.twinx()
    ax7.spines["right"].set_position(("outward", 60))
    ax7.set_ylabel("Peak Alpha Mean", color='r')
    ax7.plot(a_range, Y_Top10_peak_alpha_mean, color='r', label="Peak Alpha Mean")
    ax7.fill_between(a_range, Y_Top10_peak_alpha_mean - np.sqrt(Y_Top10_peak_alpha_var), Y_Top10_peak_alpha_mean + np.sqrt(Y_Top10_peak_alpha_var), color='r', alpha=0.2)
    #add_scatter_fit(ax7, a_range, Y_Top10_peak_alpha_mean, 'r', "Peak Alpha Mean")
    ax7.tick_params(axis='y', labelcolor='r')

    ax8 = ax5.twinx()
    ax8.spines["right"].set_position(("outward", 120))
    ax8.set_ylabel("Score", color='c')
    ax8.plot(a_range, Y_Top10_score_mean, color='c', label="Score")
    ax8.fill_between(a_range, Y_Top10_score_mean - np.sqrt(Y_Top10_score_var), Y_Top10_score_mean + np.sqrt(Y_Top10_score_var), color='c', alpha=0.2)
    #add_scatter_fit(ax8, a_range, Y_Top10_score_mean, 'c', "Score")
    ax8.tick_params(axis='y', labelcolor='c')


    ax5.grid()
    ax5.set_title("Youth Group: Multi-Axis Analysis with Fits")

    plt.tight_layout()
    plt.savefig('graphs_mean/1_to_70/Top_10_parameters_with_error.png', dpi=300, bbox_inches="tight")

def return_contour(a_val):
    step_x_low = []  # x-axis values (e.g., mean delay indices)
    step_y_low = []  # y-axis values (e.g., global coupling indices)
    
    step_x_high = []  # x-axis values (e.g., mean delay indices)
    step_y_high = []
    plot_list = np.nan * np.zeros((len(gc), len(md)))
    for r in range(43):
            for i in range(len(gc)):
                for j in range(len(md)):
                    if r-0.5 <= all_psd_peak[a_val][i][j] < r+0.5:
                        plot_list[i][j] = r
    
    
    for i in range(plot_list.shape[0]):  # Iterate over rows
        #iterate over columns in reverse order to find the first step
        for j in range(plot_list.shape[1]-1, -1, -1):  # Iterate over columns in reverse order
            value = int(plot_list[i, j])  # Get the value at this position
            value_1r = int(plot_list[i,j+1]) if j+1 < plot_list.shape[1] else 1000  # Get the next value to check for step continuity
            value_1d = int(plot_list[i-1,j]) if i-1 >= 0 else 0  # Get the previous value to check for step continuity
            value_1u = int(plot_list[i+1,j]) if i+1 < plot_list.shape[0] else 0  # Get the next value to check for step continuity
            value_1l = int(plot_list[i,j-1]) if j-1 >= 0 else 0  # Get the previous value to check for step continuity

            #text_color = cmap(norm(value))  # Get corresponding color from colormap
            
            
            
            if value < 7:  # If it's not the first step, continue to the next column

                if j == plot_list.shape[1] - 1 and value_1d >= 7:
                    step_x_low.append(j+0.5)
                    step_y_low.append(i-0.5)

                if value_1l >= 7:
                    step_x_low.append(j-0.5)  
                    step_y_low.append(i-0.5)

                if i == plot_list.shape[0] - 1 and value_1l >= 7:
                    step_x_low.append(j-0.5)
                    step_y_low.append(i+0.5)
            
            if value >13:  # If it's not the first step, continue to the next column
    
                if value_1r <= 13 and value_1u <= 13 and value_1l <= 13:  # Check if the step is not at the edge
                    step_x_high.append(j+0.5)  # Append the x position of the step
                    step_y_high.append(i-0.5)  # Append the y position of the step

                    step_x_high.append(j+0.5)  # Append the x position of the step
                    step_y_high.append(i+0.5)

                    step_x_high.append(j-0.5)  # Append the x position of the step
                    step_y_high.append(i+0.5)  # Append the y position of the step
                    
                elif value_1l <= 13 and value_1l >0:
                    step_x_high.append(j-0.5)
                    step_y_high.append(i+0.5)

                    step_x_high.append(j-0.5)  # Append the x position of the step
                    step_y_high.append(i-0.5)
                elif 13>=value_1r:
                    step_x_high.append(j+0.5)  # Append the x position of the step
                    step_y_high.append(i-0.5)  # Append the y position of the step

                if i == plot_list.shape[0] - 1:
                    step_x_high.append(j+0.5)
                    step_y_high.append(i+0.5)

                if j == plot_list.shape[1] - 1 and value_1u <= 13:
                    step_x_high.append(j+0.5)
                    step_y_high.append(i+0.5)
    return step_x_low, step_y_low, step_x_high, step_y_high


def compute_psd(signals, fs, nfft_factor=5):

    """Compute the Power Spectral Density (PSD) of the signal."""
    frequencies, psd = signal.welch(signals.real, fs=fs, nperseg=2 * fs, nfft=nfft_factor * fs)
    psd_mean         = np.mean(psd, axis=0)
    return frequencies, psd_mean

def plot_psd(max_freq=250):
    filepath = f'/vol/specs07/BasDrost/SL_simulation/simulation_results/signal_C13.97_md6_a-4.00_dt0.1.npy'
                
    # Load and preprocess the signal
    sig            = np.load(filepath)
    sigg = stat.zscore(sig, axis=-1)

    # Compute and plot the power spectral density
    fs               = 1000  # Sampling frequency in Hz
    frequencies, psd = compute_psd(sigg, fs)
    plt.figure(figsize=(7, 5))
    plt.plot(frequencies[:max_freq], 10 * np.log10(psd[:max_freq]), linewidth=3)
    plt.title('Power Spectral Density (PSD)', fontsize=16)
    plt.xlabel('Frequency (Hz)', fontsize=14)
    plt.xticks(fontsize=12)
    plt.yticks(fontsize=12)
    plt.ylabel('Power Spectral Density (dB)', fontsize=14)
    plt.grid(True)
    plt.xlim(0, 30)
    plt.axvline(x=8, color='r', linestyle='--')
    plt.axvline(x=13, color='r', linestyle='--')
    plt.tight_layout()
    plt.savefig('graphs_mean/1_to_70/PSD_poster.png', dpi=300)
                    
                    

Filter=False
contour = True
#plot_mean_correlation(Filter, contour)
#plot_max_correlation(Filter, contour)
#plot_psd_peak(Filter)
#plot_psd_parameter_relation(contour)
#plot_psd_parameter_relation_dif_color()
#plot_SEM_peak(Filter, contour)
#plot_SEM_corr(Filter, contour)
#top_10_mean_corr()
#top_10_mean_corr_all()
#analyse_top_10()
#plot_psd()
            

